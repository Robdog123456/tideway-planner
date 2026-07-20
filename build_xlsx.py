#!/usr/bin/env python3
"""Build the Tideway row-window workbook from grid.json (house palette + traffic lights)."""
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = '/Users/rob/Claude/rowing-nk/rowing-coach/_athletes/rob/tideway-row-windows-2026-07-15.xlsx'
NAVY, TEAL, GOLD = '0D1B2A', '1B7A8C', 'F2A71B'
GREEN, RED, GREY = '2E7D32', 'C62828', 'D9D9D9'
navy_f  = PatternFill('solid', fgColor=NAVY)
teal_f  = PatternFill('solid', fgColor=TEAL)
green_f = PatternFill('solid', fgColor=GREEN)
gold_f  = PatternFill('solid', fgColor=GOLD)
red_f   = PatternFill('solid', fgColor=RED)
grey_f  = PatternFill('solid', fgColor=GREY)
band_f  = PatternFill('solid', fgColor='EEF3F5')   # light teal tint zebra
white   = Font(color='FFFFFF', name='Calibri', size=11)
whiteB  = Font(color='FFFFFF', bold=True, name='Calibri', size=11)
navyB   = Font(color=NAVY, bold=True, name='Calibri', size=11)
base    = Font(name='Calibri', size=11)
ctr = Alignment(horizontal='center', vertical='center')
lft = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin = Side(style='thin', color='BFCAD0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

LFILL = {'GREEN': green_f, 'AMBER': gold_f, 'RED': red_f}
LFONT = {'GREEN': whiteB, 'AMBER': navyB, 'RED': whiteB}
OFILL = {'ROW': green_f, 'Caution': gold_f, "Don't row": red_f}
OTXT  = {'ROW': 'Row', 'Caution': 'Row (care)', "Don't row": "Don't row"}

DNAME = {'15':'Wednesday','16':'Thursday','17':'Friday','18':'Saturday',
         '19':'Sunday','20':'Monday','21':'Tuesday'}
COLS = [('Time',9), ('Height\n(m, Putney CD)',15), ('Tide status',30),
        ('Wind (BST forecast)',22), ('Tide',8), ('Wind',8), ('Verdict',12), ('Notes',40)]


def runs(rows, pred):
    out, start, last = [], None, None
    for r in rows:
        if pred(r):
            start = start or r['time']; last = r['time']
        elif start:
            out.append((start, last)); start = None
    if start: out.append((start, last))
    return out

def endslot(t):
    h, m = int(t[:2]), int(t[3:]) + 30
    return f'{h + m//60:02d}:{m%60:02d}'

def winfmt(w):
    return ', '.join(f'{a}–{endslot(b)}' for a, b in w) if w else '—'


def cell(ws, r, c, v, fill=None, font=base, align=lft, bd=True):
    x = ws.cell(r, c, v)
    if fill: x.fill = fill
    x.font = font; x.alignment = align
    if bd: x.border = border
    return x


def day_sheet(wb, key, data):
    rows = data['grid'][key]; hwlw = data['putney_hwlw'].get(key, [])
    dd = key[-2:]
    ws = wb.create_sheet(f'{DNAME[dd][:3]} {int(dd)} Jul')
    ws.sheet_view.showGridLines = False
    for i, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    cell(ws, 1, 1, f'{DNAME[dd]} {int(dd)} July 2026  —  Tideway (Putney), single scull',
         navy_f, whiteB, ctr, bd=False)
    ws.row_dimensions[1].height = 24
    # tide line
    tline = 'Putney tides (BST):   ' + '     '.join(
        f"{'HW' if e['type']=='HW' else 'LW'} {e['time']}  ({e['height_m']} m)" for e in hwlw)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    cell(ws, 2, 1, tline, teal_f, white, Alignment(horizontal='center', vertical='center'), bd=False)
    ws.row_dimensions[2].height = 20
    # header
    hr = 3
    for i, (name, _) in enumerate(COLS, 1):
        cell(ws, hr, i, name, navy_f, whiteB, ctr)
    ws.row_dimensions[hr].height = 28
    # data
    for j, r in enumerate(rows):
        rr = hr + 1 + j
        zebra = band_f if j % 2 else None
        tcell = cell(ws, rr, 1, r['time'], grey_f if 'dark' in r['notes'] else zebra, base, ctr)
        cell(ws, rr, 2, r['height_putney_m'], zebra, base, ctr)
        cell(ws, rr, 3, r['tide_status'], zebra)
        wind = f"{r['wind_mph']} mph {r['wind_dir_txt']}  (gust {r['gust_mph']})"
        cell(ws, rr, 4, wind, zebra, base, Alignment(horizontal='center', vertical='center'))
        cell(ws, rr, 5, r['tide_light'].title(), LFILL[r['tide_light']], LFONT[r['tide_light']], ctr)
        cell(ws, rr, 6, r['wind_light'].title(), LFILL[r['wind_light']], LFONT[r['wind_light']], ctr)
        cell(ws, rr, 7, OTXT[r['overall']], OFILL[r['overall']], LFONT[{'ROW':'GREEN','Caution':'AMBER',"Don't row":'RED'}[r['overall']]], ctr)
        cell(ws, rr, 8, r['notes'], zebra, Font(name='Calibri', size=10, italic=True))
    ws.freeze_panes = 'A4'


def summary_sheet(wb, data):
    ws = wb.create_sheet('Summary'); ws.sheet_view.showGridLines = False
    widths = [16, 34, 30, 26, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.merge_cells('A1:E1')
    cell(ws, 1, 1, 'Tideway single-scull row windows  —  week of Wed 15 Jul 2026 (Putney, BST)',
         navy_f, whiteB, ctr, bd=False); ws.row_dimensions[1].height = 26
    ws.merge_cells('A2:E2')
    cell(ws, 2, 1, 'Green = ideal (calm)   ·   Amber = rowable, be aware   ·   Red = don’t row.   '
                   'Verdict needs BOTH tide and wind clear. Check the live PLA/LRC flag on the day.',
         teal_f, white, Alignment(horizontal='center', vertical='center', wrap_text=True), bd=False)
    ws.row_dimensions[2].height = 28
    hdr = ['Day', 'Putney tides (BST)', 'Rowable (green + amber)', 'Ideal (calm / green)', 'Best window']
    for i, h in enumerate(hdr, 1):
        cell(ws, 3, i, h, navy_f, whiteB, ctr)
    ws.row_dimensions[3].height = 24
    for j, key in enumerate(sorted(data['grid'])):
        rows = data['grid'][key]; dd = key[-2:]; rr = 4 + j
        tides = '  '.join(f"{e['type']} {e['time']}({e['height_m']}m)" for e in data['putney_hwlw'].get(key, []))
        rowable = runs(rows, lambda r: r['overall'] != "Don't row")
        ideal = runs(rows, lambda r: r['overall'] == 'ROW')
        best = max(rowable, key=lambda w: (int(endslot(w[1])[:2]) * 60 + int(endslot(w[1])[3:])) -
                                          (int(w[0][:2]) * 60 + int(w[0][3:]))) if rowable else None
        zebra = band_f if j % 2 else None
        cell(ws, rr, 1, f'{DNAME[dd][:3]} {int(dd)} Jul', zebra, navyB, ctr)
        cell(ws, rr, 2, tides, zebra, Font(name='Calibri', size=10), ctr)
        cell(ws, rr, 3, winfmt(rowable), zebra, base, ctr)
        cell(ws, rr, 4, winfmt(ideal), zebra, base, ctr)
        cell(ws, rr, 5, winfmt([best]) if best else '—', zebra, navyB, ctr)
    ws.freeze_panes = 'A4'
    return ws


def method_sheet(wb):
    ws = wb.create_sheet('Method & caveats'); ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 96
    cell(ws, 1, 1, 'Method & caveats', navy_f, whiteB, ctr, bd=False)
    ws.merge_cells('A1:B1'); ws.row_dimensions[1].height = 22
    items = [
        ('Purpose', 'Planning aid for when a single scull can go out on the Tideway at Putney (London RC). Predictions only — the live club/PLA flag on the day always governs.'),
        ('Tide source', 'Port of London Authority predictions, London Bridge gauge (tidepredictions.pla.co.uk), converted to Putney Bridge with PLA secondary-port offsets: HW +31 min / −1.0 m, LW +98 min / −0.5 m. Heights = metres above chart datum (≈±0.3 m). Cross-checked against WillyWeather Putney (HW/LW within ≤ a few minutes).'),
        ('Wind source', 'Open-Meteo hourly forecast for Putney (51.467, −0.216), miles per hour, interpolated to 30-min slots. Confidence degrades after ~day 5.'),
        ('Tide light', 'LRC BLACK-flag rule ("do not boat at low tide") quantified as a low-water-avoidance band. RED = ~1 h before to ~2 h after low water (shallow + strong early flood). AMBER = the flanking half-hours. GREEN = the rest of the ebb (HW to ~2 h before LW) and the risen flood (~2.5 h after LW to HW) and high water.'),
        ('Wind light', 'Driven by SUSTAINED wind speed (calibrated to on-water experience: Wed 15 Jul 06:30–07:45 was fine at sustained ~10 mph with 18–21 mph gusts, NE on the ebb). GREEN ≤ 8 mph; AMBER 9–13 mph; RED > 13 mph, or gusts ≥ 32 mph, or wind-over-tide once already breezy (≥ 13 mph). Gusts are shown for information but do not close the window on their own.'),
        ('Wind-over-tide', 'Wind opposing the running stream (flood runs upstream ~WSW; ebb runs downstream ~ENE — approximate course axis). Flagged when the stream is running (not within ~45 min of the turn) and the wind opposes it; it only forces RED when the sustained wind is already ≥ 13 mph. Slack water and high water are the calmest.'),
        ('Verdict', 'ROW = tide and wind both green. Row (care) = either amber, none red. Don’t row = either red.'),
        ('Not modelled', 'The live PLA ebb-tide flag (fluvial flow after rain / lower-than-predicted tides) is set twice daily and is NOT forecastable a week out — check pla.co.uk/ebb-tide-flag-warning and the club flag on the day. Also: darkness (rowing-in-the-dark restriction; slots before sunrise are marked), commercial traffic, debris, and your own risk assessment as a Tideway Expert sculler.'),
        ('Built', 'From data pulled 15 Jul 2026. Re-run compute.py to refresh.'),
    ]
    for k, v in enumerate(items):
        rr = 2 + k
        cell(ws, rr, 1, v[0], teal_f, whiteB, Alignment(horizontal='left', vertical='top', wrap_text=True))
        cell(ws, rr, 2, v[1], None, base, lft)
        ws.row_dimensions[rr].height = 46


def main():
    data = json.load(open('grid.json'))
    wb = Workbook(); wb.remove(wb.active)
    summary_sheet(wb, data)
    for key in sorted(data['grid']):
        day_sheet(wb, key, data)
    method_sheet(wb)
    wb.save(OUT)
    print('wrote', OUT, '| sheets:', wb.sheetnames)


if __name__ == '__main__':
    main()
